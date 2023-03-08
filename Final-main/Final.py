#Riley Ringer and Javier Rodriguez Group 2
#This program is intended to make entering data into an excel table easier.
#When ran, this program will display a GUI window with entry boxes and buttons.
#The program checks to see if a "Data.xlsx" file exists if not it then creates one to store user data.
#Once the required fields have been entered you may then submit your data.
#The document can then be viewed in the same folder for easy viewing and access.


from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib
from openpyxl.styles import Font

#Root Window
root=Tk()
root.title("Customer Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg='#e9967a')


#Creates an excel workbook. First checks to see if file already exists. If not then creates workbook named 'Data.xlsx'.

file = pathlib.Path('Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] ="Name"
    sheet['B1'] ="Birth"
    sheet['C1'] ="Gender"
    sheet['D1'] ="PhoneNumber"
    sheet['E1'] ="Address"
    sheet['F1'] ="Email"

    #Changing font style to bold on column attribute
    font = Font(bold=True)
    sheet['A1'].font= font
    sheet['B1'].font= font
    sheet['C1'].font= font
    sheet['D1'].font= font
    sheet['E1'].font= font
    sheet['F1'].font= font

    file.save('Data.xlsx')



def submit():
    # Open the Excel file and selects the active sheet, in this case it's Data.xlsx
    file = openpyxl.load_workbook('Data.xlsx')
    sheet = file.active
    
    # Get the values from the text entry boxes
    name = name_value.get()
    birth = birth_value.get()
    gender = gender_combobox.get()
    phone = phone_value.get()
    address = address_entry.get("1.0",'end-1c')
    email = email_entry.get()

    # Validate that all fields are filled out
    if not all([name, birth, gender, phone, address, email]):
        messagebox.showerror("Error", "All fields are required.")
        return
    
    # Add the values to the next available row in the sheet
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = birth
    sheet.cell(row=next_row, column=3).value = gender
    sheet.cell(row=next_row, column=4).value = phone
    sheet.cell(row=next_row, column=5).value = address
    sheet.cell(row=next_row, column=6).value = email

    #clearing entry and value boxes
    name_value.set('')
    phone_value.set('')
    birth_value.set('')
    address_entry.delete('1.0',END)
    email_value.set('')
    
    # Save the Excel file
    file.save('Data.xlsx')
    
    # Show a message box to indicate success
    messagebox.showinfo("Success", "User information submitted successfully.")




#Defining the clear function. Clear all values
def clear():
    name_value.set('')
    phone_value.set('')
    birth_value.set('')
    address_entry.delete('1.0',END)
    email_value.set('')
    


#Adding Icon
icon=PhotoImage(file='logo.png')
root.iconphoto(False,icon)

#Heading in window
Label(root,text='Please Enter Customer Data',font='times 13 bold',bg='#e9967a',fg='black').place(x=20,y=20)

#Label Creation
Label(root,text='Name',font='23',bg='#e9967a',fg='black').place(x=50,y=100)
Label(root,text='Phone No.',font='23',bg='#e9967a',fg='black').place(x=50,y=150)
Label(root,text='Date of Birth',font='23',bg='#e9967a',fg='black').place(x=50,y=200)
Label(root,text='Email',font='23',bg='#e9967a',fg='black').place(x=375,y=200)
Label(root,text='Address',font='23',bg='#e9967a',fg='black').place(x=50,y=250)
Label(root,text='Gender',font='23',bg='#e9967a',fg='black').place(x=375,y=250)

#Entry Values
name_value = StringVar()
phone_value = StringVar()
birth_value = StringVar()
email_value = StringVar()

name_entry = Entry(root, textvariable=name_value,width=40,bd=2,font=20)
phone_entry = Entry(root, textvariable=phone_value,width=40,bd=2,font=20)
birth_entry = Entry(root, textvariable=birth_value,width=15,bd=2,font=20)
email_entry = Entry(root,textvariable=email_value,width=15,bd=2,font=20)

#Gender Combobox
gender_combobox = Combobox(root,values=['Male','Female','Other'],font='arial 14', state='r',width=14)
gender_combobox.place(x=450,y=250)

#Address text box
address_entry = Text(root,bd=2,width=20,height=2)

#placing entry boxes
name_entry.place(x=200,y=100)
phone_entry.place(x=200,y=150)
birth_entry.place(x=200,y=200)
address_entry.place(x=200,y=250)
email_entry.place(x=450,y=200)

#Button Creation
Button(root,text='Submit',bg='#e9967a',fg='white',width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text='Clear',bg='#e9967a',fg='white',width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text='Exit',bg='#e9967a',fg='white',width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)


root.mainloop()